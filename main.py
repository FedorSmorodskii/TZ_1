"""
В программе показана минимальная информация с сайтов конкурентов. При необходимости, мы можем увеличить воронку сбора.
Так же рекомендуется сделать программу асинхронной, это даст дикий буст в скорости.
Мистрал я выбрал по причине их бесплатных ключей API (у меня их порядка 80шт, что может повысить скорость при
многопоточности до максимума).
"""

from bs4 import BeautifulSoup
import time
from fake_useragent import UserAgent
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import random
import requests


# Настройки
COMPETITOR_SITES = [
    'https://el-dent.ru',
    'https://stomatorg.ru',
    'https://www.nika-dent.ru',
    'https://aveldent.ru',
    'https://stomdevice.ru'
]
mistral_keys = [
    'u5Q5Tky8dKA0LX7lqq9hIZrXfe1FWFqq',
    '8us4O062l89asi3OdoLWav0TewMpoBms',
    '56jCtyCEtr0vfk7dA8biQZCzl3up7Wn4',
]
OUTPUT_EXCEL = 'dental_products_descriptions.xlsx'


# Парсинг данных с сайта конкурента
def parse_competitor_product(url):
    headers = {'User-Agent': UserAgent().random}
    try:
        response = requests.get(url, headers=headers, timeout=10)
        soup = BeautifulSoup(response.text, 'html.parser')

        title = soup.find('h1').text.strip() if soup.find('h1') else ''

        description = ''
        possible_desc = soup.find_all('div', class_=lambda x: x and 'description' in x.lower())
        for desc in possible_desc:
            description += desc.text.strip() + '\n\n'

        return {
            'original_title': title,
            'original_description': description.strip(),
            'url': url
        }
    except Exception as e:
        print(f"Ошибка при парсинге {url}: {e}")
        return None


def generate_with_mistral(prompt):
    retries = 5
    API_URL = "https://api.mistral.ai/v1/chat/completions"

    data = {
        "model": "mistral-large-2407",
        "messages": [{"role": "user", "content": f"{prompt}"}],
        "temperature": 0.5,
        "max_tokens": 1000,
        "top_p": 1,
        "presence_penalty": 0.5,
        "frequency_penalty": 0.5,
    }

    for attempt in range(1, retries + 1):
        key = random.choice(mistral_keys)
        headers = {
            'Authorization': f'Bearer {key}',
            'Content-Type': 'application/json'
        }

        try:
            response = requests.post(API_URL, json=data, headers=headers)
            response.raise_for_status()  # Raises HTTPError for bad responses

            result = response.json()
            return result.get("choices", [{}])[0].get("message", {}).get("content", "")

        except requests.exceptions.HTTPError as e:
            print(f'HTTP error: {e} (attempt {attempt}/{retries})')
            if attempt == retries:
                return f"HTTP error: {str(e)}"

        except requests.exceptions.RequestException as e:
            print(f'Request failed: {e} (attempt {attempt}/{retries})')
            if attempt == retries:
                return f"Request failed: {str(e)}"

        time.sleep(1)  # Add delay between retries

    return "Не удалось получить ответ"


# Генерация уникального описания
def generate_unique_content(product_data):
    prompt = f"""
    [INST] На основе следующего описания стоматологического товара создай уникальный SEO-оптимизированный текст:
    Название товара: {product_data['original_title']}
    Описание: {product_data['original_description']}

    Требования:
    1. Текст должен быть уникальным и не копировать исходный
    2. Использовать профессиональную терминологию
    3. Добавить преимущества товара
    4. Оптимизировать для SEO
    5. Верстка по правилам Dental First:
       - Абзацы разделяются двойным переносом строки
       - Списки оформляются через дефисы
       - Важные моменты выделяются **жирным**
    [/INST]
    """

    result = generate_with_mistral(prompt)
    return result.split('[/INST]')[-1].strip() if result else None


# Генерация мета-данных
def generate_meta_data(product_title, product_description):
    prompt = f"""
    [INST] На основе информации о товаре создай SEO-метаданные в следующем формате:
    Title: [здесь title до 60 символов]
    Description: [здесь description до 160 символов]
    Keywords: [здесь 5-10 ключевых слов через запятую]

    Информация о товаре:
    Название: {product_title}
    Описание: {product_description}
    [/INST]
    """

    result = generate_with_mistral(prompt)
    # print(result)
    if not result:
        return None

    # Удаляем все до [/INST], если есть
    response_part = result.split('[/INST]')[-1].strip()

    # Инициализируем метаданные
    meta = {'title': '', 'description': '', 'keywords': ''}

    # Парсим строки
    lines = [line.strip() for line in response_part.split('\n') if line.strip()]

    for line in lines:
        if line.lower().startswith('title:'):
            meta['title'] = line.split(':', 1)[1].strip()
        elif line.lower().startswith('description:'):
            meta['description'] = line.split(':', 1)[1].strip()
        elif line.lower().startswith('keywords:'):
            meta['keywords'] = line.split(':', 1)[1].strip()

    return meta


# Создание и форматирование Excel-файла
def create_excel_file(products_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Описания товаров"

    # Заголовки
    headers = [
        'Название товара',
        'Описание товара',
        'Title',
        'Description',
        'Keywords',
        'URL источника'
    ]
    ws.append(headers)

    # Стили для заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    # Применение стилей к заголовкам
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = border

    # Добавление данных
    for product in products_data:
        ws.append([
            product['name'],
            product['description'],
            product['meta']['title'],
            product['meta']['description'],
            product['meta']['keywords'],
            product['url']
        ])

    # Настройка ширины столбцов и переноса текста
    column_widths = {
        'A': 35,  # Название товара
        'B': 80,  # Описание товара
        'C': 60,  # Title
        'D': 80,  # Description
        'E': 40,  # Keywords
        'F': 30  # URL источника
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Форматирование ячеек с данными
    for row in ws.iter_rows(min_row=2, max_row=len(products_data) + 1, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = border

    # Сохранение файла
    wb.save(OUTPUT_EXCEL)
    print(f"Excel-файл успешно сохранен: {OUTPUT_EXCEL}")


# Основная функция
def main():
    start_time = time.time()

    # Тестовые товары
    test_products = [
        {'name': 'All-Bond 3 part A and B (2х6мл) - адгезив двойного отверждения БИСКО В-36010Р',
         'url': 'https://el-dent.ru/id/all-bond3-adgeziv-bisko.html'},
    ]

    products_data = []

    # Обработка товаров
    for product in test_products:
        print(f"Обработка товара: {product['name']}")

        # Парсинг данных
        product_data = parse_competitor_product(product['url'])
        if not product_data:
            continue

        # Генерация описания
        unique_description = generate_unique_content(product_data)
        if not unique_description:
            continue

        # Генерация мета-данных
        meta = generate_meta_data(product['name'], unique_description)
        if not meta:
            continue

        # Сохранение данных
        products_data.append({
            'name': product['name'],
            'description': unique_description,
            'meta': meta,
            'url': product['url']
        })

        # Пауза между запросами
        time.sleep(random.randint(1, 3))

    # Создание Excel-файла
    create_excel_file(products_data)

    end_time = time.time()
    total_time = end_time - start_time
    print(f"Завершено! Общее время выполнения: {total_time // 60} минут {total_time % 60:.2f} секунд")


if __name__ == "__main__":
    main()